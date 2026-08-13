"""Excel 库存导入、正式商品库匹配与批量打包导出。"""

from __future__ import annotations

import io
import json
import re
import unicodedata
import urllib.request
import zipfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote, urlparse

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from server.auth_util import require_perms, write_op_log
from server.config import settings
from server.db import get_conn, next_id, rows_as_dicts
from server.schemas import ApiOk

router = APIRouter(prefix="/api/excel", tags=["excel"])

APPROVAL_HEADERS = {"国药准字", "批准文号", "药品批准文号"}
SPEC_HEADERS = {"规格", "规格/型号", "规格型号"}
NAME_HEADERS = ("品名", "商品名称", "通用名称", "药品名称")
MANUFACTURER_HEADERS = ("生产厂家", "生产企业", "厂家", "上市许可持有人")
# 搜索词必须优先使用具体商品名；库存表通常把泛化的“通用名称”放在更靠前的列。
SEARCH_HEADERS = ("商品名称", "品名", "标题", "通用名称")
IMPORT_TEMPLATE_HEADERS = ["国药准字", "品名", "规格", "生产厂家"]
EXPORT_HEADERS = [
    "商品ID",
    "商品名称",
    "品名",
    "规格",
    "国药准字",
    "品牌",
    "生产厂家",
    "goodsId",
    "goodsId列表价",
    "goodsId多规格售价",
    "售价区间",
    "商品主图",
    "匹配状态",
]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_key(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).lower()


def _normalize_approval(value: Any) -> str:
    """统一准字号中的全半角、大小写和空白差异。"""
    text = unicodedata.normalize("NFKC", _text(value)).upper()
    return re.sub(r"\s+", "", text)


def _normalize_spec(value: Any) -> str:
    """统一规格写法，并忽略资料库附加的最外层包装单位。"""
    text = unicodedata.normalize("NFKC", _text(value)).lower()
    text = text.replace("×", "*").replace("／", "/")
    text = re.sub(r"\s+", "", text)
    # S/s 是库存系统常用的最小包装数量简写；与片、粒、丸、支等按数量等价。
    text = re.sub(r"(?<=\d)(?:片|粒|丸|支|袋|贴|枚|只)", "s", text)
    # 库内常见“10g*9袋/盒”或“12粒/瓶/盒”，库存表会省略最外层包装单位。
    return re.sub(r"(?:/(?:盒|瓶|袋|支|板|包|罐|桶))+$", "", text)


def _normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", text)


def _normalize_manufacturer(value: Any) -> str:
    text = _normalize_name(value)
    return re.sub(r"(?:有限责任公司|股份有限公司|有限公司|制药厂|药厂)$", "", text)


def _text_field_matches(actual: Any, expected: Any, normalizer) -> bool:
    actual_key, expected_key = normalizer(actual), normalizer(expected)
    return bool(actual_key and expected_key and (actual_key in expected_key or expected_key in actual_key))


def _find_header(headers: Iterable[Any], aliases: set[str]) -> int | None:
    normalized = {_header_key(v) for v in aliases}
    for index, value in enumerate(headers):
        if _header_key(value) in normalized:
            return index
    return None


def _find_header_by_priority(headers: Iterable[Any], aliases: Iterable[str]) -> int | None:
    """按业务优先级找列，而不是按 Excel 中从左到右的列顺序。"""
    header_keys = [_header_key(value) for value in headers]
    for alias in aliases:
        key = _header_key(alias)
        if key in header_keys:
            return header_keys.index(key)
    return None


def _search_keyword(name: Any, spec: Any) -> str:
    """精确商品名 + 规格能显著减少同通用名下错误厂家/包装商品。"""
    clean_name = re.sub(r"^[^\w\u4e00-\u9fff]+", "", _text(name))
    parts = [clean_name, _text(spec)]
    return re.sub(r"\s+", " ", " ".join(part for part in parts if part)).strip()[:256]


def _read_excel_rows(filename: str, content: bytes) -> list[tuple[Any, ...]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise ValueError("仅支持 .xlsx/.xls 文件")
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - deployment dependency guard
            raise ValueError("服务端缺少 xlrd，暂时不能读取 .xls 文件") from exc
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        return [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def _money(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        number = Decimal(str(value))
    except Exception:
        return _text(value)
    rendered = f"{number:.2f}".rstrip("0").rstrip(".")
    return f"¥{rendered}"


def _price_range(minimum: Any, maximum: Any) -> str:
    if minimum is None or maximum is None:
        return ""
    return f"{_money(minimum)}-{_money(maximum)}"


def _image_url(value: Any) -> str:
    image = _text(value).replace("\\", "/")
    if not image or image.startswith(("http://", "https://", "/")):
        return image
    return f"/media/{image.lstrip('/')}"


def _candidate(cur, hit: dict[str, Any], platform_code: str, range_min: Any, range_max: Any) -> dict[str, Any]:
    cur.execute(
        """
        SELECT SPEC, SALE_PRICE
          FROM T_GOODS_LIBRARY
         WHERE PLATFORM_CODE = :platform_code AND GOODS_ID = :goods_id
         ORDER BY UPDATE_TIME DESC NULLS LAST, LIBRARY_ID DESC
        """,
        {"platform_code": platform_code, "goods_id": hit.get("goods_id")},
    )
    spec_prices: list[str] = []
    seen: set[tuple[str, str]] = set()
    for variant in rows_as_dicts(cur):
        spec = _text(variant.get("spec"))
        price = _money(variant.get("sale_price"))
        key = (spec, price)
        if key in seen:
            continue
        seen.add(key)
        spec_prices.append(" ".join(x for x in (spec, price) if x))

    return {
        "product_id": hit.get("library_id"),
        "goods_id": _text(hit.get("goods_id")),
        "product_name": _text(hit.get("product_name")),
        "sell_name": _text(hit.get("sell_name")),
        "spec": _text(hit.get("spec")),
        "approval_no": _text(hit.get("approval_no")),
        "brand": _text(hit.get("brand")),
        "manufacturer": _text(hit.get("manufacturer")),
        "list_price": hit.get("list_price"),
        "multi_spec_prices": "、".join(spec_prices),
        "price_range": _price_range(range_min, range_max),
        "main_image": _image_url(hit.get("main_image")),
        "update_time": hit.get("update_time"),
    }


def _match_one(
    cur, platform_code: str, approval_no: str, product_name: str, spec: str, manufacturer: str,
) -> list[dict[str, Any]]:
    approval_key = _normalize_approval(approval_no)
    spec_key = _normalize_spec(spec)
    if not approval_key or not spec_key:
        return []
    cur.execute(
        """
        SELECT LIBRARY_ID, GOODS_ID, PRODUCT_NAME, SELL_NAME, SPEC, APPROVAL_NO,
               BRAND, MANUFACTURER, LIST_PRICE, SALE_PRICE, MAIN_IMAGE, UPDATE_TIME
         FROM T_GOODS_LIBRARY
         WHERE PLATFORM_CODE = :platform_code
           AND REPLACE(UPPER(TRIM(APPROVAL_NO)), ' ', '') = :approval_no
         ORDER BY UPDATE_TIME DESC NULLS LAST, LIBRARY_ID DESC
        """,
        {"platform_code": platform_code, "approval_no": approval_key},
    )
    hits = [
        hit for hit in rows_as_dicts(cur)
        if _normalize_spec(hit.get("spec")) == spec_key
        and (
            _text_field_matches(hit.get("product_name"), product_name, _normalize_name)
            or _text_field_matches(hit.get("sell_name"), product_name, _normalize_name)
        )
        and _text_field_matches(hit.get("manufacturer"), manufacturer, _normalize_manufacturer)
    ]
    if not hits:
        return []
    prices = [hit.get("sale_price") for hit in hits if hit.get("sale_price") is not None]
    range_min = min(prices) if prices else None
    range_max = max(prices) if prices else None
    return [_candidate(cur, hit, platform_code, range_min, range_max) for hit in hits]


def _result_row(
    row_index: int,
    approval_no: str,
    spec: str,
    product_name: str,
    manufacturer: str,
    candidates: list[dict[str, Any]],
    search_keyword: str = "",
    original_row: dict[str, str] | None = None,
) -> dict[str, Any]:
    if not candidates:
        return {
            "row_index": row_index,
            "matched": False,
            "match_status": "unmatched",
            "match_count": 0,
            "input_approval_no": approval_no,
            "input_spec": spec,
            "input_product_name": product_name,
            "input_manufacturer": manufacturer,
            "original_row": original_row or {},
            "search_keyword": search_keyword,
            "product_id": None,
            "goods_id": "",
            "product_name": "",
            "sell_name": "",
            "spec": spec,
            "approval_no": approval_no,
            "brand": "",
            "manufacturer": "",
            "list_price": None,
            "multi_spec_prices": "",
            "price_range": "",
            "main_image": "",
            "candidates": [],
        }
    status = "multiple" if len(candidates) > 1 else "unique"
    return {
        "row_index": row_index,
        "matched": True,
        "match_status": status,
        "match_count": len(candidates),
        "input_approval_no": approval_no,
        "input_spec": spec,
        "input_product_name": product_name,
        "input_manufacturer": manufacturer,
        "original_row": original_row or {},
        "search_keyword": search_keyword,
        **candidates[0],
        "candidates": candidates if len(candidates) > 1 else [],
    }


@router.get("/template")
def download_template(_=Depends(require_perms("excel:import"))):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"
    sheet.append(IMPORT_TEMPLATE_HEADERS)
    sheet.append(["国药准字H32020475", "盐酸氨溴索口服溶液", "20ml*2支", "某某制药有限公司"])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="409EFF")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 32
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=goods_library_match_template.xlsx"},
    )


@router.post("/match")
async def match_excel(
    file: UploadFile = File(...),
    platform_code: str = "pinduoduo",
    user=Depends(require_perms("excel:match")),
):
    try:
        rows = _read_excel_rows(file.filename or "", await file.read())
    except Exception as exc:
        return ApiOk(ok=False, message=f"Excel 读取失败：{exc}")
    if not rows:
        return ApiOk(ok=False, message="空文件")

    approval_index = _find_header(rows[0], APPROVAL_HEADERS)
    spec_index = _find_header(rows[0], SPEC_HEADERS)
    name_index = _find_header_by_priority(rows[0], NAME_HEADERS)
    manufacturer_index = _find_header_by_priority(rows[0], MANUFACTURER_HEADERS)
    search_index = _find_header_by_priority(rows[0], SEARCH_HEADERS)
    missing = []
    if approval_index is None:
        missing.append("国药准字/批准文号")
    if spec_index is None:
        missing.append("规格/规格型号")
    if name_index is None:
        missing.append("品名/商品名称")
    if manufacturer_index is None:
        missing.append("生产厂家/生产企业")
    if missing:
        return ApiOk(ok=False, message=f"缺少必填列：{'、'.join(missing)}")

    results: list[dict[str, Any]] = []
    unique_count = multiple_count = unmatched_count = 0
    with get_conn() as conn:
        cur = conn.cursor()
        for excel_row, row in enumerate(rows[1:], start=2):
            approval_no = _text(row[approval_index]) if approval_index < len(row) else ""
            spec = _text(row[spec_index]) if spec_index < len(row) else ""
            product_name = _text(row[name_index]) if name_index is not None and name_index < len(row) else ""
            manufacturer = _text(row[manufacturer_index]) if manufacturer_index is not None and manufacturer_index < len(row) else ""
            search_name = _text(row[search_index]) if search_index is not None and search_index < len(row) else product_name
            search_keyword = _search_keyword(search_name, spec)
            original_row = {
                _text(header) or f"列{index + 1}": _text(row[index]) if index < len(row) else ""
                for index, header in enumerate(rows[0])
            }
            if not approval_no and not spec:
                continue
            candidates = _match_one(cur, platform_code, approval_no, product_name, spec, manufacturer) if all((approval_no, product_name, spec, manufacturer)) else []
            result = _result_row(excel_row, approval_no, spec, product_name, manufacturer, candidates, search_keyword, original_row)
            results.append(result)
            if result["match_status"] == "unique":
                unique_count += 1
            elif result["match_status"] == "multiple":
                multiple_count += 1
            else:
                unmatched_count += 1
        write_op_log(
            cur,
            user_id=user["user_id"],
            username=user["username"],
            action="goods_library_excel_match",
            module="excel",
            detail=f"平台={platform_code} 唯一={unique_count} 多匹配={multiple_count} 未匹配={unmatched_count}",
        )

    return ApiOk(
        data={
            "total": len(results),
            "matched": unique_count + multiple_count,
            "unique": unique_count,
            "multiple": multiple_count,
            "unmatched": unmatched_count,
            "rows": results,
        }
    )


def _export_values(row: dict[str, Any]) -> list[Any]:
    status = {"unique": "唯一匹配", "multiple": "多匹配项", "unmatched": "未匹配"}.get(
        row.get("match_status"), "未匹配"
    )
    return [
        row.get("product_id"),
        row.get("product_name"),
        row.get("sell_name"),
        row.get("spec"),
        row.get("approval_no"),
        row.get("brand"),
        row.get("manufacturer"),
        row.get("goods_id"),
        row.get("list_price"),
        row.get("multi_spec_prices"),
        row.get("price_range"),
        row.get("main_image"),
        status,
    ]


def _workbook_bytes(row: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品数据"
    sheet.append(EXPORT_HEADERS)
    sheet.append(_export_values(row))
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="409EFF")
        cell.alignment = Alignment(horizontal="center")
    for column in sheet.columns:
        letter = column[0].column_letter
        longest = max(len(_text(cell.value)) for cell in column)
        sheet.column_dimensions[letter].width = min(max(longest + 2, 12), 36)
    sheet["I2"].number_format = '¥0.00'
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _safe_name(value: Any, fallback: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", _text(value)).strip(" .")
    return (name or fallback)[:80]


def _local_media_path(image: str) -> Path | None:
    if not image.startswith("/media/"):
        return None
    root = Path(settings.image_dir).resolve()
    candidate = (root / image.removeprefix("/media/")).resolve()
    try:
        candidate.relative_to(root)
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def _image_bytes(image: Any) -> tuple[bytes, str] | None:
    image_url = _text(image)
    local = _local_media_path(image_url)
    if local:
        return local.read_bytes(), local.suffix.lower() or ".jpg"
    if not image_url.startswith(("http://", "https://")):
        return None
    request = urllib.request.Request(image_url, headers={"User-Agent": "SJZQ-ExcelExport/1.0"})
    with urllib.request.urlopen(request, timeout=10) as response:
        data = response.read(10 * 1024 * 1024 + 1)
        if len(data) > 10 * 1024 * 1024:
            return None
        content_type = response.headers.get_content_type()
    suffix = Path(urlparse(image_url).path).suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        suffix = {"image/png": ".png", "image/webp": ".webp", "image/gif": ".gif"}.get(content_type, ".jpg")
    return data, suffix


@router.post("/export-batch")
def export_batch(body: dict, user=Depends(require_perms("excel:export"))):
    rows = [row for row in (body.get("rows") or []) if row.get("matched") and row.get("product_id")]
    if not rows:
        return ApiOk(ok=False, message="请至少勾选一条已匹配商品")
    platform_code = _text(body.get("platform_code")) or "pinduoduo"
    platform_name = platform_code
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT PLATFORM_NAME FROM SJZQ_PLATFORM WHERE PLATFORM_CODE = :code", {"code": platform_code})
        result = cur.fetchone()
        if result and result[0]:
            platform_name = _text(result[0])
        write_op_log(
            cur,
            user_id=user["user_id"],
            username=user["username"],
            action="goods_library_excel_export",
            module="excel",
            detail=f"平台={platform_code} 导出={len(rows)}",
        )

    output = io.BytesIO()
    used_folders: set[str] = set()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for index, row in enumerate(rows, start=1):
            base = _safe_name(f"{row.get('product_id')}_{row.get('product_name')}", f"商品_{index}")
            folder = base
            suffix_index = 2
            while folder in used_folders:
                folder = f"{base}_{suffix_index}"
                suffix_index += 1
            used_folders.add(folder)
            archive.writestr(f"{folder}/{folder}.xlsx", _workbook_bytes(row))
            try:
                image = _image_bytes(row.get("main_image"))
            except Exception:
                image = None
            if image:
                archive.writestr(f"{folder}/商品主图{image[1]}", image[0])
    output.seek(0)

    filename = f"{_safe_name(platform_name, platform_code)}_{datetime.now():%Y%m%d_%H%M%S}.zip"
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/export-matched")
def export_matched(body: dict, _=Depends(require_perms("excel:export"))):
    """兼容旧前端：把全部结果导出为单个 Excel。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "匹配结果"
    sheet.append(EXPORT_HEADERS)
    for row in body.get("rows") or []:
        sheet.append(_export_values(row))
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=goods_library_matched.xlsx"},
    )


@router.post("/unmatched-to-task")
def unmatched_to_task(body: dict, user=Depends(require_perms("task:dispatch"))):
    """把 Excel 未匹配行下发给 Android，由 APP 按准字+规格核对详情。"""
    rows = [row for row in (body.get("rows") or []) if not row.get("matched")]
    targets = []
    for row in rows:
        approval = _text(row.get("input_approval_no") or row.get("approval_no"))[:128]
        name = _text(row.get("input_product_name") or row.get("product_name"))[:256]
        spec = _text(row.get("input_spec") or row.get("spec"))[:256]
        manufacturer = _text(row.get("input_manufacturer") or row.get("manufacturer"))[:256]
        keyword = _text(row.get("search_keyword") or name or approval)[:256]
        original_row = row.get("original_row") or {
            "国药准字": approval, "品名": name, "规格": spec, "生产厂家": manufacturer,
        }
        if approval and name and spec and manufacturer and keyword:
            targets.append({
                "keyword": keyword, "approval": approval, "name": name,
                "spec": spec, "manufacturer": manufacturer,
                "original_row": original_row,
            })
    if not targets:
        return ApiOk(ok=False, message="没有可生成任务的未匹配条目")
    platform = body.get("platform_code") or "pinduoduo"
    try:
        device_id = int(body.get("device_id"))
    except (TypeError, ValueError):
        return ApiOk(ok=False, message="请选择有效的采集设备")
    if device_id <= 0:
        return ApiOk(ok=False, message="请选择有效的采集设备")
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT PLATFORM_CODE, OWNER_USER_ID FROM SJZQ_DEVICE WHERE DEVICE_ID = :device_id",
            {"device_id": device_id},
        )
        device = cur.fetchone()
        if not device:
            return ApiOk(ok=False, message="所选采集设备不存在")
        if _text(device[0]) != _text(platform):
            return ApiOk(ok=False, message="所选设备与当前平台不一致")
        if user.get("role_code") != "super_admin" and int(device[1] or 0) != int(user["user_id"]):
            return ApiOk(ok=False, message="运营只能向本人绑定的设备下发任务")
        task_id = next_id(cur, "SJZQ_SEQ_TASK")
        cur.execute(
            """
            INSERT INTO SJZQ_TASK (
                TASK_ID, TASK_NAME, TASK_TYPE, PLATFORM_CODE, DEVICE_ID, STATUS, PRIORITY,
                KEYWORD_TEXT, TARGET_COUNT, CONFIG_JSON, CREATE_USER_ID, CREATE_USERNAME,
                REVIEW_STATUS
            ) VALUES (
                :task_id, :task_name, 'collect', :platform, :device_id, 'pending', 5,
                :keywords, :target_count, :config_json, :user_id, :username, 'pending'
            )
            """,
            {
                "task_id": task_id,
                "task_name": body.get("task_name") or f"Excel未匹配补采-{task_id}",
                "platform": platform,
                "device_id": device_id,
                "keywords": "\n".join(target["keyword"] for target in targets),
                "target_count": len(targets),
                "config_json": json.dumps(
                    {
                        "match_mode": "approval_name_spec_manufacturer",
                        "max_detail": max(1, min(int(body.get("max_detail") or 10), 30)),
                    },
                    ensure_ascii=False,
                ),
                "user_id": user["user_id"],
                "username": user["username"],
            },
        )
        for index, target in enumerate(targets):
            item_id = next_id(cur, "SJZQ_SEQ_TASK_ITEM")
            cur.execute(
                """
                INSERT INTO SJZQ_TASK_ITEM (
                    ITEM_ID, TASK_ID, ROW_INDEX, KEYWORD,
                    TARGET_SPEC, TARGET_APPROVAL, TARGET_NAME, TARGET_MANUFACTURER,
                    ORIGINAL_ROW_JSON, STATUS
                ) VALUES (
                    :item_id, :task_id, :row_index, :keyword,
                    :target_spec, :target_approval, :target_name, :target_manufacturer,
                    :original_row_json, 'pending'
                )
                """,
                {
                    "item_id": item_id,
                    "task_id": task_id,
                    "row_index": index,
                    "keyword": target["keyword"],
                    "target_spec": target["spec"],
                    "target_approval": target["approval"],
                    "target_name": target["name"],
                    "target_manufacturer": target["manufacturer"],
                    "original_row_json": json.dumps(target["original_row"], ensure_ascii=False),
                },
            )
        write_op_log(
            cur,
            user_id=user["user_id"],
            username=user["username"],
            action="excel_android_match_task",
            module="excel",
            detail=f"未匹配下发 Android 任务 #{task_id} 设备={device_id} 共 {len(targets)} 条",
        )
        return ApiOk(data={"task_id": task_id, "device_id": device_id, "count": len(targets)})
